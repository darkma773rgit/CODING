from __future__ import annotations

from typing import List

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference


def autosize_columns(ws_name_to_widths: dict) -> None:
    for ws, widths in ws_name_to_widths.items():
        for col_index, width in widths.items():
            ws.column_dimensions[get_column_letter(col_index)].width = width


def style_header_row(ws, header_row: int) -> None:
    header_fill = PatternFill(start_color="FFDDEBF7", end_color="FFDDEBF7", fill_type="solid")
    header_font = Font(bold=True, color="FF1F4E79")
    thin = Side(border_style="thin", color="FFBFBFBF")
    border = Border(top=thin, left=thin, right=thin, bottom=thin)
    for cell in ws[header_row]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border


def style_table_region(ws, start_row: int, end_row: int, start_col: int, end_col: int) -> None:
    thin = Side(border_style="thin", color="FFDDDDDD")
    border = Border(top=thin, left=thin, right=thin, bottom=thin)
    for row in ws.iter_rows(min_row=start_row, max_row=end_row, min_col=start_col, max_col=end_col):
        for cell in row:
            cell.border = border
            if cell.column > start_col:  # numeric area
                cell.number_format = "_$#,##0.00_);[Red](_$#,##0.00)"
                cell.alignment = Alignment(horizontal="right")


def create_monthly_sheet(ws, title: str, months: List[str], data_rows: int = 20) -> None:
    ws.title = title
    ws.freeze_panes = "B2"

    headers = ["Category", *months, "Total"]
    ws.append(headers)
    style_header_row(ws, 1)

    # Pre-fill empty rows with running row totals
    for r in range(2, 2 + data_rows):
        ws.cell(row=r, column=1, value="")
        # Blank numeric cells for months
        for c in range(2, 2 + len(months)):
            ws.cell(row=r, column=c, value=None)
        # Row total = SUM across month columns
        start_letter = get_column_letter(2)
        end_letter = get_column_letter(1 + len(months))
        ws.cell(row=r, column=2 + len(months), value=f"=SUM({start_letter}{r}:{end_letter}{r})")

    # Grand total row
    total_row = 2 + data_rows + 1
    ws.cell(row=total_row, column=1, value="Grand Total").font = Font(bold=True)

    # Monthly column totals
    for c in range(2, 2 + len(months)):
        col_letter = get_column_letter(c)
        ws.cell(row=total_row, column=c, value=f"=SUM({col_letter}2:{col_letter}{total_row-1})")

    # Total of totals column
    last_col = 2 + len(months)
    last_letter = get_column_letter(last_col)
    ws.cell(row=total_row, column=last_col, value=f"=SUM({last_letter}2:{last_letter}{total_row-1})")
    ws.cell(row=total_row, column=last_col).font = Font(bold=True)

    # Styling for table region
    style_table_region(ws, start_row=1, end_row=total_row, start_col=1, end_col=last_col)

    # Column widths
    widths = {1: 22}
    for idx in range(2, 2 + len(months)):
        widths[idx] = 12
    widths[2 + len(months)] = 14
    autosize_columns({ws: widths})


def create_summary_sheet(wb: Workbook, months: List[str], income_ws_name: str, expenses_ws_name: str, data_rows: int) -> None:
    ws = wb.create_sheet("Summary", 0)
    ws.freeze_panes = "B3"

    headers = ["Category", *months, "Annual Total"]
    ws.append(["Budget Summary"])
    ws["A1"].font = Font(size=14, bold=True, color="FF1F4E79")
    ws.append(headers)
    style_header_row(ws, 2)

    # Helper to sum a month across a sheet's data rows
    def month_sum_formula(sheet_name: str, month_col_index: int) -> str:
        col_letter = get_column_letter(month_col_index)
        start_row = 2
        end_row = 1 + data_rows
        return f"=SUM('{sheet_name}'!{col_letter}{start_row}:'{sheet_name}'!{col_letter}{end_row})"

    # Income row
    income_row = 3
    ws.cell(row=income_row, column=1, value="Income").font = Font(bold=True)
    for mi, _m in enumerate(months, start=2):
        ws.cell(row=income_row, column=mi, value=month_sum_formula(income_ws_name, mi))
    ws.cell(row=income_row, column=2 + len(months), value=f"=SUM(B{income_row}:{get_column_letter(1+len(months))}{income_row})")

    # Expenses row
    expenses_row = 4
    ws.cell(row=expenses_row, column=1, value="Expenses").font = Font(bold=True)
    for mi, _m in enumerate(months, start=2):
        ws.cell(row=expenses_row, column=mi, value=month_sum_formula(expenses_ws_name, mi))
    ws.cell(row=expenses_row, column=2 + len(months), value=f"=SUM(B{expenses_row}:{get_column_letter(1+len(months))}{expenses_row})")

    # Net row
    net_row = 5
    ws.cell(row=net_row, column=1, value="Net (Income - Expenses)").font = Font(bold=True, color="FF0A7F00")
    for mi, _m in enumerate(months, start=2):
        inc = f"{get_column_letter(mi)}{income_row}"
        exp = f"{get_column_letter(mi)}{expenses_row}"
        ws.cell(row=net_row, column=mi, value=f"={inc}-{exp}")
    ws.cell(row=net_row, column=2 + len(months), value=f"=SUM(B{net_row}:{get_column_letter(1+len(months))}{net_row})")

    # Number formats and borders for the data area (Budget block)
    style_table_region(ws, start_row=2, end_row=5, start_col=1, end_col=2 + len(months))

    # Actuals block
    actuals_income_row = 7
    actuals_expenses_row = 8
    actuals_net_row = 9
    ws.cell(row=actuals_income_row, column=1, value="Actual Income").font = Font(bold=True)
    ws.cell(row=actuals_expenses_row, column=1, value="Actual Expenses").font = Font(bold=True)
    ws.cell(row=actuals_net_row, column=1, value="Actual Net (Income - Expenses)").font = Font(bold=True, color="FF0A7F00")

    income_actuals_name = f"{income_ws_name} Actuals"
    expenses_actuals_name = f"{expenses_ws_name} Actuals"

    for mi, _m in enumerate(months, start=2):
        # Monthly sums from Actuals sheets
        ws.cell(row=actuals_income_row, column=mi, value=month_sum_formula(income_actuals_name, mi))
        ws.cell(row=actuals_expenses_row, column=mi, value=month_sum_formula(expenses_actuals_name, mi))
        inc_cell = f"{get_column_letter(mi)}{actuals_income_row}"
        exp_cell = f"{get_column_letter(mi)}{actuals_expenses_row}"
        ws.cell(row=actuals_net_row, column=mi, value=f"={inc_cell}-{exp_cell}")

    # Annual totals for Actuals block
    ws.cell(row=actuals_income_row, column=2 + len(months), value=f"=SUM(B{actuals_income_row}:{get_column_letter(1+len(months))}{actuals_income_row})")
    ws.cell(row=actuals_expenses_row, column=2 + len(months), value=f"=SUM(B{actuals_expenses_row}:{get_column_letter(1+len(months))}{actuals_expenses_row})")
    ws.cell(row=actuals_net_row, column=2 + len(months), value=f"=SUM(B{actuals_net_row}:{get_column_letter(1+len(months))}{actuals_net_row})")

    # Borders and number formats for Actuals block
    style_table_region(ws, start_row=7, end_row=9, start_col=1, end_col=2 + len(months))

    # Variance block (Actual - Budget)
    variance_row = 11
    ws.cell(row=variance_row, column=1, value="Variance (Actual Net - Budget Net)").font = Font(bold=True, color="FF9C0006")
    for mi, _m in enumerate(months, start=2):
        budget_net = f"{get_column_letter(mi)}{net_row}"
        actual_net = f"{get_column_letter(mi)}{actuals_net_row}"
        ws.cell(row=variance_row, column=mi, value=f"={actual_net}-{budget_net}")
    ws.cell(row=variance_row, column=2 + len(months), value=f"=SUM(B{variance_row}:{get_column_letter(1+len(months))}{variance_row})")

    style_table_region(ws, start_row=11, end_row=11, start_col=1, end_col=2 + len(months))

    # Notes / instructions
    ws["A13"] = (
        "Instructions: Enter your categories on the Budget (Income/Expenses) and Actuals sheets. "
        "Use rows 2-21 for line items; totals auto-calculate here."
    )
    ws["A13"].alignment = Alignment(wrap_text=True)

    # Column widths
    widths = {1: 28}
    for idx in range(2, 2 + len(months)):
        widths[idx] = 14
    widths[2 + len(months)] = 16
    autosize_columns({ws: widths})

    # Chart: Budget vs Actual Net by Month
    chart = BarChart()
    chart.title = "Budget vs Actual Net"
    chart.y_axis.title = "Amount"
    chart.x_axis.title = "Month"

    # Categories are month headers in row 2, columns B..(B+months)
    categories = Reference(ws, min_col=2, min_row=2, max_col=1 + len(months), max_row=2)
    budget_net_series = Reference(ws, min_col=2, min_row=net_row, max_col=1 + len(months), max_row=net_row)
    actual_net_series = Reference(ws, min_col=2, min_row=actuals_net_row, max_col=1 + len(months), max_row=actuals_net_row)

    chart.add_data(budget_net_series, titles_from_data=False, from_rows=True)
    chart.add_data(actual_net_series, titles_from_data=False, from_rows=True)
    chart.set_categories(categories)
    chart.width = 20
    chart.height = 10
    ws.add_chart(chart, "A15")


def build_workbook() -> Workbook:
    months = [
        "Jan",
        "Feb",
        "Mar",
        "Apr",
        "May",
        "Jun",
        "Jul",
        "Aug",
        "Sep",
        "Oct",
        "Nov",
        "Dec",
    ]

    wb = Workbook()

    create_summary_sheet(wb, months, income_ws_name="Income", expenses_ws_name="Expenses", data_rows=20)
    income_ws = wb.create_sheet("Income")
    expenses_ws = wb.create_sheet("Expenses")
    income_actuals_ws = wb.create_sheet("Income Actuals")
    expenses_actuals_ws = wb.create_sheet("Expenses Actuals")

    create_monthly_sheet(income_ws, "Income", months, data_rows=20)
    create_monthly_sheet(expenses_ws, "Expenses", months, data_rows=20)
    create_monthly_sheet(income_actuals_ws, "Income Actuals", months, data_rows=20)
    create_monthly_sheet(expenses_actuals_ws, "Expenses Actuals", months, data_rows=20)

    # Move Summary to be the first sheet
    wb._sheets.sort(key=lambda s: 0 if s.title == "Summary" else 1)
    return wb


def main() -> None:
    wb = build_workbook()
    output_name = "Budget_Template.xlsx"
    wb.save(output_name)
    print(f"Created {output_name}")


if __name__ == "__main__":
    main()


